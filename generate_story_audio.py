# -*- coding: utf-8 -*-
"""
根据Excel文件中的故事内容生成音频文件
使用edge-tts生成中文语音
"""
import os
import sys
import re
from openpyxl import load_workbook
from openpyxl.cell.rich_text import TextBlock, CellRichText
import asyncio
import edge_tts

# 配置
OUTPUT_DIR = "generated_audio"  # 音频文件输出目录
VOICE = "zh-CN-XiaoxiaoNeural"  # 中文女声，可选: zh-CN-YunxiNeural(男声), zh-CN-XiaoxiaoNeural(女声)

async def generate_audio(text: str, output_file: str, voice: str = VOICE):
    """生成音频文件"""
    try:
        print(f"正在生成音频: {output_file}")
        print(f"文本长度: {len(text)} 字符")
        
        # 使用edge-tts生成音频
        communicate = edge_tts.Communicate(text, voice)
        await communicate.save(output_file)
        
        # 检查文件是否生成成功
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"[OK] 音频生成成功: {output_file} ({file_size / 1024:.2f} KB)")
            return True
        else:
            print(f"[FAIL] 音频生成失败: {output_file}")
            return False
    except Exception as e:
        print(f"[ERROR] 生成音频时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

def clean_filename(title: str) -> str:
    """清理文件名，移除特殊字符"""
    if not title:
        return "untitled"
    
    # 移除换行符、制表符等空白字符
    filename = title.replace("\n", "").replace("\r", "").replace("\t", " ").strip()
    
    # 移除或替换特殊字符
    filename = filename.replace(":", "_").replace("/", "_").replace("\\", "_") \
                       .replace("?", "_").replace("*", "_").replace("\"", "_") \
                       .replace("<", "_").replace(">", "_").replace("|", "_") \
                       .replace("：", "_").replace(""", "_").replace(""", "_") \
                       .replace("'", "_").replace("'", "_")
    
    # 移除连续的空格和下划线
    import re
    filename = re.sub(r'[_\s]+', '_', filename)
    filename = filename.strip('_')
    
    # 限制文件名长度（Windows限制255字符，但为了安全限制到200）
    if len(filename) > 200:
        filename = filename[:200]
    
    # 如果文件名为空，使用默认名称
    if not filename:
        filename = "untitled"
    
    return filename

def extract_content_text(content_parts):
    """从内容部分提取纯文本（移除加粗标记）"""
    if isinstance(content_parts, list):
        text = ""
        for part in content_parts:
            text += part.get("text", "")
        return text
    else:
        # 移除**加粗标记**
        text = str(content_parts) if content_parts else ""
        text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)  # 移除**标记但保留文本
        return text

async def main():
    """主函数"""
    # 创建输出目录
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"创建输出目录: {OUTPUT_DIR}")
    
    # 读取Excel文件
    print("正在读取Excel文件...")
    wb = load_workbook('Story_v2.xlsx', rich_text=True)
    ws = wb.active
    
    # 获取列名
    headers = [cell.value for cell in ws[1]]
    print(f"列名: {headers}")
    
    # 读取所有故事数据
    stories = []
    for row_idx in range(2, ws.max_row + 1):
        story = {}
        
        # 读取每一列
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            
            # 如果是内容列，需要处理加粗格式
            if header == "内容" and value:
                content_parts = []
                
                # 检查是否是富文本
                if isinstance(value, CellRichText):
                    # 处理富文本
                    for block in value:
                        if isinstance(block, TextBlock):
                            text = block.text
                            is_bold = False
                            if block.font and hasattr(block.font, 'b') and block.font.b:
                                is_bold = True
                            content_parts.append({"text": text, "bold": is_bold})
                        else:
                            # 普通文本块
                            text = str(block)
                            content_parts.append({"text": text, "bold": False})
                else:
                    # 普通文本，检查单元格字体
                    text = str(value) if value else ""
                    is_bold = False
                    if cell.font and hasattr(cell.font, 'b') and cell.font.b:
                        is_bold = True
                    content_parts.append({"text": text, "bold": is_bold})
                
                story[header] = content_parts
            else:
                story[header] = value
        
        # 只添加非空故事
        if any(v for v in story.values() if v):
            stories.append(story)
    
    print(f"\n成功读取 {len(stories)} 条故事")
    
    # 处理日期和标题 - 根据实际Excel列顺序调整
    # Excel列顺序: 日期、字数统计、题目、内容
    date_col = headers[0] if len(headers) > 0 else None
    title_col = None
    content_col = None
    
    # 查找标题和内容列
    for i, h in enumerate(headers):
        if h and ("题目" in str(h) or "标题" in str(h) or "title" in str(h).lower()):
            title_col = h
        if h and ("内容" in str(h) or "content" in str(h).lower()):
            content_col = h
    
    if not title_col:
        title_col = headers[2] if len(headers) > 2 else headers[1]
    if not content_col:
        content_col = headers[3] if len(headers) > 3 else headers[2]
    
    print(f"日期列: {date_col}, 标题列: {title_col}, 内容列: {content_col}")
    
    # 生成音频文件
    success_count = 0
    fail_count = 0
    
    print(f"\n开始生成音频文件（使用音色: {VOICE}）...")
    print("=" * 80)
    
    for i, story in enumerate(stories, 1):
        date_str = story.get(date_col, "")
        title = story.get(title_col, "")
        content_parts = story.get(content_col, [])
        
        # 处理日期格式
        if isinstance(date_str, (int, float)):
            day = int(date_str)
            date_formatted = f"2025-01-{day:02d}"
        elif isinstance(date_str, str) and "月" in date_str:
            match = re.search(r'(\d+)月', date_str)
            if match:
                day = int(match.group(1))
                date_formatted = f"2025-01-{day:02d}"
            else:
                date_formatted = f"2025-01-{i:02d}"
        else:
            date_formatted = f"2025-01-{i:02d}"
        
        if not title:
            print(f"\n[{i}/{len(stories)}] 跳过：标题为空")
            continue
        
        # 提取内容文本
        content_text = extract_content_text(content_parts)
        if not content_text or len(content_text.strip()) < 10:
            print(f"\n[{i}/{len(stories)}] 跳过：内容为空或太短")
            continue
        
        # 生成文件名
        clean_title = clean_filename(title)
        output_file = os.path.join(OUTPUT_DIR, f"{clean_title}.mp3")
        
        print(f"\n[{i}/{len(stories)}] {date_formatted} - {title}")
        print(f"  内容长度: {len(content_text)} 字符")
        
        # 生成音频
        success = await generate_audio(content_text, output_file, VOICE)
        
        if success:
            success_count += 1
        else:
            fail_count += 1
        
        # 添加延迟，避免请求过快
        await asyncio.sleep(0.5)
    
    print("\n" + "=" * 80)
    print(f"音频生成完成！")
    print(f"成功: {success_count} 个")
    print(f"失败: {fail_count} 个")
    print(f"输出目录: {os.path.abspath(OUTPUT_DIR)}")
    print(f"\n提示: 将生成的音频文件复制到 app/src/main/assets/story_audio/ 目录下")

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n\n用户中断")
    except Exception as e:
        print(f"\n\n错误: {e}")
        import traceback
        traceback.print_exc()
