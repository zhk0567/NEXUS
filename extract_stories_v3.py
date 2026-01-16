# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.cell.rich_text import TextBlock, CellRichText
import re

# 读取Excel文件
wb = load_workbook('Story_v2.xlsx', rich_text=True)
ws = wb.active

# 获取列名
headers = [cell.value for cell in ws[1]]
print(f"列名: {headers}")

# 读取所有故事数据
stories = []
for row_idx in range(2, ws.max_row + 1):
    story = {}
    
    # 读取每一列
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        
        # 如果是内容列，需要处理加粗格式
        if header == "内容" and value:
            content_parts = []
            
            # 检查是否是富文本
            if isinstance(value, CellRichText):
                # 处理富文本
                for block in value:
                    if isinstance(block, TextBlock):
                        text = block.text
                        is_bold = False
                        if block.font and hasattr(block.font, 'b') and block.font.b:
                            is_bold = True
                        content_parts.append({"text": text, "bold": is_bold})
                    else:
                        # 普通文本块
                        text = str(block)
                        content_parts.append({"text": text, "bold": False})
            else:
                # 普通文本，检查单元格字体
                text = str(value) if value else ""
                is_bold = False
                if cell.font and hasattr(cell.font, 'b') and cell.font.b:
                    is_bold = True
                content_parts.append({"text": text, "bold": is_bold})
            
            story[header] = content_parts
        else:
            story[header] = value
    
    # 只添加非空故事
    if any(v for v in story.values() if v):
        stories.append(story)

print(f"\n成功读取 {len(stories)} 条故事")

# 生成Kotlin代码，使用**标记加粗
kotlin_code = []
kotlin_code.append("    private fun initializeStories() {")

for i, story in enumerate(stories, 1):
    # 获取日期、标题、内容
    date_col = headers[0]
    title_col = headers[2] if len(headers) > 2 else headers[1]
    content_col = headers[3] if len(headers) > 3 else headers[2]
    
    date_str = story.get(date_col, "")
    title = story.get(title_col, "")
    content_parts = story.get(content_col, [])
    
    # 处理日期格式
    if isinstance(date_str, (int, float)):
        day = int(date_str)
        date_formatted = f"2025-01-{day:02d}"
    elif isinstance(date_str, str) and "月" in date_str:
        # 提取月份数字
        match = re.search(r'(\d+)月', date_str)
        if match:
            day = int(match.group(1))
            date_formatted = f"2025-01-{day:02d}"
        else:
            date_formatted = f"2025-01-{i:02d}"
    else:
        date_formatted = f"2025-01-{i:02d}"
    
    # 构建内容字符串，使用**标记加粗
    content_str = ""
    if isinstance(content_parts, list):
        for part in content_parts:
            text = part.get("text", "")
            is_bold = part.get("bold", False)
            
            if is_bold:
                content_str += f"**{text}**"
            else:
                content_str += text
    else:
        content_str = str(content_parts) if content_parts else ""
    
    # 转义特殊字符
    title = title.replace("$", "\\$").replace("\"", "\\\"")
    
    # 生成Kotlin代码
    kotlin_code.append(f'        addStory("{date_formatted}", "{title}", """')
    # 将内容按行分割，每行添加适当的缩进
    lines = content_str.split('\n')
    for line in lines:
        # 转义特殊字符
        line = line.replace("$", "\\$")
        kotlin_code.append(f"            {line}")
    kotlin_code.append('        """.trimIndent(), "温馨故事")')
    kotlin_code.append("")

kotlin_code.append("    }")

# 保存到文件
with open('stories_kotlin.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(kotlin_code))

print(f"\nKotlin代码已生成，共 {len(stories)} 条故事")
print("代码已保存到 stories_kotlin.txt")
