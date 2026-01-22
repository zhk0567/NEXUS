#!/usr/bin/env python3
"""
数据库管理器
"""
import pymysql
import hashlib
import uuid
import json
import logging
import time
import threading
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from database_config import DATABASE_CONFIG, CREATE_TABLES_SQL, INIT_DATABASE_SQL, DEFAULT_ADMIN

logger = logging.getLogger(__name__)

class DatabaseManager:
    """数据库管理器 - 优化版本（线程安全）"""
    
    def __init__(self):
        self.connection = None
        self.connection_pool = []
        self.max_connections = 5
        # 线程锁保护连接访问
        self._lock = threading.Lock()
        # 性能优化：添加查询缓存
        self.query_cache = {}
        self.cache_ttl = 300  # 5分钟缓存
        # 连接重试配置
        self.max_retries = 3
        self.retry_delay = 1
        self.init_database()
    
    def connect(self, use_database=True):
        """连接到数据库"""
        try:
            # 添加连接参数以提高稳定性
            config = DATABASE_CONFIG.copy()
            # 如果数据库不存在，先不指定数据库连接
            if not use_database:
                config.pop('database', None)
            
            config.update({
                'autocommit': True,
                'charset': 'utf8mb4',
                'use_unicode': True,
                'connect_timeout': 10,
                'read_timeout': 60,  # 增加读取超时
                'write_timeout': 60,  # 增加写入超时
                'init_command': "SET SESSION sql_mode='STRICT_TRANS_TABLES', wait_timeout=28800, interactive_timeout=28800",
                'sql_mode': 'STRICT_TRANS_TABLES',
                'cursorclass': pymysql.cursors.DictCursor
            })
            self.connection = pymysql.connect(**config)
            # 设置连接保持活跃，定期ping保持连接
            self.connection.ping(reconnect=True)
            # 设置自动重连
            self.connection.autocommit(True)
            # 数据库连接成功，不输出日志
        except Exception as e:
            logger.error(f"❌ 数据库连接失败: {e}")
            raise
    
    def reconnect(self):
        """重新连接数据库（线程安全）"""
        try:
            if self.connection:
                try:
                    self.connection.close()
                except:
                    pass  # 忽略关闭时的错误
                self.connection = None
            self.connect()
            # 数据库重新连接成功，不输出日志
        except Exception as e:
            logger.error(f"❌ 数据库重新连接失败: {e}")
            raise
    
    def _get_fresh_connection(self):
        """获取新的数据库连接"""
        config = DATABASE_CONFIG.copy()
        config.update({
            'autocommit': True,
            'charset': 'utf8mb4',
            'use_unicode': True,
            'connect_timeout': 10,
            'read_timeout': 60,  # 增加读取超时
            'write_timeout': 60,  # 增加写入超时
            'init_command': "SET SESSION wait_timeout=28800, interactive_timeout=28800",
            'cursorclass': pymysql.cursors.DictCursor
        })
        return pymysql.connect(**config)
    
    def is_connection_healthy(self):
        """检查数据库连接是否健康"""
        try:
            if not self.connection or not self.connection.open:
                return False
            # 执行简单查询测试连接
            self.connection.ping(reconnect=False)
            return True
        except:
            return False
    
    def init_database(self):
        """初始化数据库和表"""
        temp_connection = None
        try:
            # 先连接到MySQL服务器（不指定数据库）
            config_no_db = DATABASE_CONFIG.copy()
            config_no_db.pop('database', None)
            config_no_db.update({
                'autocommit': True,
                'charset': 'utf8mb4',
                'use_unicode': True,
                'connect_timeout': 10,
                'read_timeout': 30,
                'write_timeout': 30,
                'cursorclass': pymysql.cursors.DictCursor
            })
            temp_connection = pymysql.connect(**config_no_db)
            
            with temp_connection.cursor() as cursor:
                # 创建数据库
                cursor.execute(INIT_DATABASE_SQL)
                temp_connection.commit()
            
            # 关闭临时连接
            temp_connection.close()
            temp_connection = None
            
            # 现在连接到指定的数据库
            self.connect(use_database=True)
            
            # 创建表
            with self.connection.cursor() as cursor:
                for table_name, sql in CREATE_TABLES_SQL.items():
                    cursor.execute(sql)
                    # 表创建成功，不输出日志
                
                self.connection.commit()
                # 数据库初始化完成，不输出日志
                
                # 创建默认管理员用户
                self.create_default_admin()
                
                # 创建默认用户（user01-user10）
                self.create_default_users()
                
        except Exception as e:
            if temp_connection:
                try:
                    temp_connection.close()
                except:
                    pass
            logger.error(f"❌ 数据库初始化失败: {e}")
            raise
    
    def create_default_admin(self):
        """创建默认管理员用户"""
        try:
            # 检查是否已存在管理员用户
            if self.get_user_by_username('admin'):
                # 管理员用户已存在，不输出日志
                return
            
            # 创建管理员用户
            self.create_user(
                user_id=DEFAULT_ADMIN['user_id'],
                username=DEFAULT_ADMIN['username'],
                password=DEFAULT_ADMIN['password'],
                is_active=DEFAULT_ADMIN['is_active']
            )
            # 默认管理员用户创建成功，不输出日志
            
        except Exception as e:
            logger.error(f"❌ 创建默认管理员用户失败: {e}")
    
    def create_default_users(self):
        """创建默认用户（user01-user10）"""
        try:
            default_users = [
                ('user01', '123456'),
                ('user02', '123456'),
                ('user03', '123456'),
                ('user04', '123456'),
                ('user05', '123456'),
                ('user06', '123456'),
                ('user07', '123456'),
                ('user08', '123456'),
                ('user09', '123456'),
                ('user10', '123456'),
            ]
            
            created_count = 0
            for username, password in default_users:
                # 检查用户是否已存在
                if self.get_user_by_username(username):
                    continue
                
                # 生成用户ID
                import uuid
                user_id = f"user_{uuid.uuid4().hex[:8]}"
                
                # 创建用户
                if self.create_user(
                    user_id=user_id,
                    username=username,
                    password=password,
                    is_active=True
                ):
                    created_count += 1
                    logger.info(f"✅ 创建默认用户: {username} (ID: {user_id})")
            
            if created_count > 0:
                logger.info(f"✅ 已创建 {created_count} 个默认用户")
            
        except Exception as e:
            logger.error(f"❌ 创建默认用户失败: {e}")
    
    def hash_password(self, password: str) -> str:
        """密码哈希"""
        return hashlib.sha256(password.encode()).hexdigest()
    
    def verify_password(self, password: str, password_hash: str) -> bool:
        """验证密码"""
        return self.hash_password(password) == password_hash
    
    def execute_with_retry(self, operation, *args, **kwargs):
        """带重试机制的数据库操作（线程安全）"""
        for attempt in range(self.max_retries):
            try:
                with self._lock:
                    # 检查连接是否有效
                    if not self.connection or not self.connection.open:
                        logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{self.max_retries})")
                        self.reconnect()
                    else:
                        # 测试连接是否真的可用，使用ping自动重连
                        try:
                            self.connection.ping(reconnect=True)
                        except:
                            logger.warning(f"⚠️ 数据库连接测试失败，尝试重新连接 (尝试 {attempt + 1}/{self.max_retries})")
                            self.reconnect()
                    
                    # 在锁内执行操作，确保线程安全
                    return operation(*args, **kwargs)
                
            except (pymysql.OperationalError, pymysql.InterfaceError, pymysql.ProgrammingError, 
                    pymysql.Error, ConnectionError, OSError) as e:
                logger.error(f"❌ 数据库操作失败 (尝试 {attempt + 1}/{self.max_retries}): {e}")
                if attempt < self.max_retries - 1:
                    try:
                        with self._lock:
                            self.reconnect()
                    except:
                        pass  # 重连失败，继续重试
                    time.sleep(self.retry_delay * (attempt + 1))  # 递增延迟
                else:
                    raise
            except Exception as e:
                logger.error(f"❌ 数据库操作失败 (尝试 {attempt + 1}/{self.max_retries}): {e}")
                if attempt < self.max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    raise
    
    def user_exists(self, user_id: str) -> bool:
        """检查用户是否存在（使用独立连接以避免并发冲突）"""
        max_retries = 3
        connection = None
        for attempt in range(max_retries):
            try:
                # 使用独立连接避免并发冲突
                connection = self._get_fresh_connection()
                
                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    # 先尝试按user_id查找，如果没找到再按username查找
                    sql = "SELECT COUNT(*) as count FROM users WHERE user_id = %s OR username = %s"
                    cursor.execute(sql, (user_id, user_id))
                    result = cursor.fetchone()
                    exists = result['count'] > 0 if result else False
        
                if connection:
                    connection.close()
                return exists
                    
            except Exception as e:
                logger.error(f"❌ 检查用户存在失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return False
        return False
    
    def create_user(self, user_id: str, username: str, password: str, email: str = None, is_active: bool = True) -> bool:
        """创建用户"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # 检查连接是否有效
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()
                
                with self.connection.cursor() as cursor:
                    # 对密码进行哈希处理
                    password_hash = self.hash_password(password)
                    
                    sql = """
                    INSERT INTO users (user_id, username, password_hash, is_active)
                    VALUES (%s, %s, %s, %s)
                    """
                    cursor.execute(sql, (user_id, username, password_hash, is_active))
                    self.connection.commit()
                    # 用户创建成功，不输出日志
                    return True
                    
            except Exception as e:
                logger.error(f"❌ 创建用户失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    try:
                        self.reconnect()
                        # 重新连接成功，不输出日志
                    except Exception as reconnect_error:
                        logger.error(f"❌ 重新连接失败: {reconnect_error}")
                else:
                    logger.error(f"❌ 创建用户最终失败，已重试 {max_retries} 次")
                    return False
    
    
    def get_user_by_username(self, username: str) -> Optional[Dict]:
        """根据用户名获取用户信息（使用独立连接）"""
        max_retries = 3
        connection = None
        for attempt in range(max_retries):
            try:
                # 使用独立连接避免并发冲突
                connection = self._get_fresh_connection()
                
                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    sql = "SELECT * FROM users WHERE username = %s"
                    cursor.execute(sql, (username,))
                    result = cursor.fetchone()
                    
                if connection:
                    connection.close()
                return result
                    
            except Exception as e:
                # 只记录有意义的错误
                if str(e) and str(e) != "(0, '')":
                    logger.error(f"❌ 获取用户信息失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return None
        return None
    
    def get_all_users(self, limit: int = 1000) -> List[Dict]:
        """获取所有用户"""
        max_retries = 3
        for attempt in range(max_retries):
            connection = None
            try:
                # 使用新的连接避免连接状态问题
                connection = self._get_fresh_connection()
                
                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    sql = "SELECT * FROM users ORDER BY created_at DESC LIMIT %s"
                    cursor.execute(sql, (limit,))
                    return cursor.fetchall()
                    
            except Exception as e:
                logger.error(f"❌ 获取所有用户失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return []
            finally:
                if connection:
                    connection.close()
        return []
    
    def get_all_reading_progress(self, limit: int = 100) -> List[Dict]:
        """获取所有阅读进度"""
        max_retries = 3
        for attempt in range(max_retries):
            connection = None
            try:
                # 使用新的连接避免连接状态问题
                connection = self._get_fresh_connection()
                
                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    sql = """
                    SELECT rp.*, u.username 
                    FROM reading_progress rp 
                    LEFT JOIN users u ON rp.user_id = u.user_id 
                    ORDER BY rp.last_read_time DESC 
                    LIMIT %s
                    """
                    cursor.execute(sql, (limit,))
                    return cursor.fetchall()
                    
            except Exception as e:
                logger.error(f"❌ 获取所有阅读进度失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return []
            finally:
                if connection:
                    connection.close()
        return []
    
    def get_user_details(self, user_id: str) -> Optional[Dict]:
        """获取用户详细信息"""
        def _get_details():
            with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                # 获取用户基本信息
                sql = "SELECT * FROM users WHERE user_id = %s"
                cursor.execute(sql, (user_id,))
                user = cursor.fetchone()
                
                if not user:
                    return None
                
                # 获取用户阅读统计
                sql = """
                SELECT 
                    COUNT(*) as total_stories,
                    SUM(CASE WHEN is_completed = 1 THEN 1 ELSE 0 END) as completed_stories,
                    AVG(reading_progress) as avg_progress,
                    MAX(last_read_time) as last_activity
                FROM reading_progress 
                WHERE user_id = %s
                """
                cursor.execute(sql, (user_id,))
                stats = cursor.fetchone()
                
                # 获取最近阅读记录
                sql = """
                SELECT story_title, reading_progress, is_completed, last_read_time
                FROM reading_progress 
                WHERE user_id = %s 
                ORDER BY last_read_time DESC 
                LIMIT 5
                """
                cursor.execute(sql, (user_id,))
                recent_reading = cursor.fetchall()
                
                return {
                    'user': user,
                    'stats': stats,
                    'recent_reading': recent_reading
                }
        return self.execute_with_retry(_get_details)
    
    def authenticate_user(self, username: str, password: str) -> Optional[Dict]:
        """用户认证 - 允许所有数据库中的激活用户登录"""
        try:
            user = self.get_user_by_username(username)
            if user and self.verify_password(password, user['password_hash']):
                # 检查用户是否激活
                if not user.get('is_active', False):
                    logger.warning(f"⚠️ 拒绝登录：用户 '{username}' 已被禁用")
                    return None
                # 更新最后登录时间
                self.update_user_login_time(user['user_id'])
                return user
            return None
        except Exception as e:
            logger.error(f"❌ 用户认证失败: {e}")
            return None
    
    def update_user_login_time(self, user_id: str):
        """更新用户登录时间（使用独立连接）"""
        connection = None
        try:
            # 使用独立连接避免并发冲突
            connection = self._get_fresh_connection()
            
            with connection.cursor() as cursor:
                sql = "UPDATE users SET last_login_at = NOW() WHERE user_id = %s"
                cursor.execute(sql, (user_id,))
                connection.commit()
                
            if connection:
                connection.close()
        except Exception as e:
            # 静默失败，不影响登录流程
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    # 移除update_user_logout_time函数 - 不再需要登出时间字段
    
    def create_session(self, user_id: str, app_type: str = 'unknown', device_info: str = None, ip_address: str = None) -> str:
        """创建用户会话（更新到 users 表）"""
        connection = None
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # 使用独立连接避免并发冲突
                connection = self._get_fresh_connection()
                
                session_id = str(uuid.uuid4())
                with connection.cursor() as cursor:
                    # 更新 users 表的会话信息
                    sql = """
                    UPDATE users 
                    SET session_id = %s,
                        app_type = %s,
                        device_info = %s,
                        ip_address = %s,
                        last_login_at = NOW()
                    WHERE user_id = %s
                    """
                    cursor.execute(sql, (session_id, app_type, device_info, ip_address, user_id))
                    connection.commit()
                    
                if connection:
                    connection.close()
                # 会话创建成功，不输出日志
                return session_id
                    
            except Exception as e:
                # 只记录有意义的错误
                if str(e) and str(e) != "(0, '')":
                    logger.error(f"❌ 创建会话失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return None
        return None
    
    def end_user_sessions(self, user_id: str, app_type: str = None) -> int:
        """结束用户的所有会话（或指定app类型的会话）（带连接健康检查和自动重连）"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # 检查连接健康状态
                if not self.is_connection_healthy():
                    logger.debug(f"数据库连接不健康，尝试重连 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()
                
                with self.connection.cursor() as cursor:
                    if app_type:
                        # 清除指定app类型的会话信息
                        sql = """
                        UPDATE users 
                        SET session_id = NULL, app_type = NULL, device_info = NULL, ip_address = NULL
                        WHERE user_id = %s AND app_type = %s
                        """
                        cursor.execute(sql, (user_id, app_type))
                    else:
                        # 清除所有会话信息
                        sql = """
                        UPDATE users 
                        SET session_id = NULL, app_type = NULL, device_info = NULL, ip_address = NULL
                        WHERE user_id = %s
                        """
                        cursor.execute(sql, (user_id,))
                    self.connection.commit()
                    return cursor.rowcount
            except (pymysql.Error, ConnectionError, OSError) as e:
                error_msg = str(e)
                # 检查是否是连接相关错误
                if any(keyword in error_msg.lower() for keyword in ['gone away', 'lost connection', 'broken pipe', 'connection aborted', '10053']):
                    logger.debug(f"数据库连接错误，尝试重连 (尝试 {attempt + 1}/{max_retries}): {error_msg}")
                    if attempt < max_retries - 1:
                        try:
                            self.reconnect()
                            time.sleep(self.retry_delay * (attempt + 1))
                            continue
                        except Exception as reconnect_error:
                            logger.error(f"重连失败: {reconnect_error}")
                    else:
                        logger.error(f"❌ 结束会话失败（连接错误，已重试 {max_retries} 次）: {error_msg}")
                        return 0
                else:
                    # 其他类型的错误，直接返回
                    if attempt == max_retries - 1:
                        logger.error(f"❌ 结束会话失败: {error_msg}")
                    return 0
            except Exception as e:
                # 其他未知错误
                if attempt == max_retries - 1:
                    logger.error(f"❌ 结束会话失败: {e}")
                return 0
        
        return 0
    
    def get_active_sessions(self, user_id: str, app_type: str = None) -> List[Dict]:
        """获取用户的活跃会话（从 users 表获取）"""
        max_retries = 3
        connection = None
        for attempt in range(max_retries):
            try:
                # 使用独立连接避免并发冲突
                connection = self._get_fresh_connection()
                
                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    if app_type:
                        sql = """
                        SELECT session_id, app_type, device_info, ip_address, last_login_at as login_time
                        FROM users
                        WHERE user_id = %s AND app_type = %s AND session_id IS NOT NULL
                        """
                        cursor.execute(sql, (user_id, app_type))
                    else:
                        sql = """
                        SELECT session_id, app_type, device_info, ip_address, last_login_at as login_time
                        FROM users
                        WHERE user_id = %s AND session_id IS NOT NULL
                        """
                        cursor.execute(sql, (user_id,))
                    results = cursor.fetchall()
                    
                    if connection:
                        connection.close()
                    return results
            except Exception as e:
                # 只记录有意义的错误（排除空错误）
                error_msg = str(e)
                if error_msg and error_msg != "(0, '')":
                    logger.error(f"❌ 获取活跃会话失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return []
        return []
    
    def get_or_create_active_session(self, user_id: str, reuse_recent: bool = False, timeout_minutes: int = 30) -> str:
        """
        获取或创建session
        
        Args:
            user_id: 用户ID
            reuse_recent: 是否复用最近的session（如果距离上次交互时间在timeout_minutes内）
            timeout_minutes: 如果启用reuse_recent，距离上次交互超过此时间则创建新session
        
        Returns:
            session_id
        """
        # 如果不需要复用，直接创建新session（每个新对话都是新session）
        if not reuse_recent:
            # 创建新对话session，不输出日志
            return self.create_session(user_id)
        
        # 如果需要复用，检查最近的交互时间
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()
                
                with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    # 查找用户最近一次交互的session和时间
                    sql = """
                    SELECT i.session_id, i.timestamp
                    FROM interactions i
                    WHERE i.user_id = %s 
                      AND i.session_id IS NOT NULL 
                      AND i.session_id != ''
                    ORDER BY i.timestamp DESC
                    LIMIT 1
                    """
                    cursor.execute(sql, (user_id,))
                    result = cursor.fetchone()
                    
                    if result and result['session_id']:
                        # 检查距离上次交互的时间
                        last_timestamp = result['timestamp']
                        # 如果timestamp是datetime对象，直接使用；如果是字符串，转换为datetime
                        if isinstance(last_timestamp, str):
                            try:
                                last_timestamp = datetime.fromisoformat(last_timestamp.replace('Z', '+00:00'))
                            except:
                                # 如果解析失败，创建新session
                                logger.warning(f"⚠️ 无法解析timestamp，创建新session")
                                return self.create_session(user_id)
                        
                        time_diff = (datetime.now() - last_timestamp).total_seconds() / 60  # 转换为分钟
                        
                        if time_diff <= timeout_minutes:
                            # 在超时时间内，复用最近的session
                            session_id = result['session_id']
                            # 复用最近的session，不输出日志
                            return session_id
                        else:
                            # 超过超时时间，创建新session
                            # 距离上次交互超过超时时间，创建新session，不输出日志
                            return self.create_session(user_id)
                    
                    # 如果没有找到任何交互记录，创建新session
                    # 未找到历史交互记录，创建新session，不输出日志
                    return self.create_session(user_id)
                    
            except Exception as e:
                logger.error(f"❌ 获取或创建session失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    try:
                        self.reconnect()
                    except:
                        pass
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    # 如果获取失败，直接创建一个新session
                    logger.warning(f"⚠️ 获取session失败，创建新session")
                    return self.create_session(user_id)
        
        # 最终失败时创建新session
        return self.create_session(user_id)
    
    def validate_or_create_session(self, user_id: str, session_id: str, timeout_minutes: int = 5) -> str:
        """
        验证session是否有效，如果无效或超时则创建新session
        
        Args:
            user_id: 用户ID
            session_id: 要验证的session_id
            timeout_minutes: 如果session最后一次交互超过此时间，创建新session
        
        Returns:
            有效的session_id（可能是原来的，也可能是新创建的）
        """
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()
                
                with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    # 查找该session的最后一次交互时间
                    sql = """
                    SELECT timestamp
                    FROM interactions 
                    WHERE user_id = %s AND session_id = %s
                    ORDER BY timestamp DESC
                    LIMIT 1
                    """
                    cursor.execute(sql, (user_id, session_id))
                    result = cursor.fetchone()
                    
                    if not result or not result.get('timestamp'):
                        # session不存在或没有任何交互记录，创建新session
                        # session不存在或无交互记录，创建新session，不输出日志
                        return self.create_session(user_id)
                    
                    # 检查距离上次交互的时间
                    last_timestamp = result['timestamp']
                    if isinstance(last_timestamp, str):
                        try:
                            last_timestamp = datetime.fromisoformat(last_timestamp.replace('Z', '+00:00'))
                        except:
                            logger.warning(f"⚠️ 无法解析timestamp，创建新session")
                            return self.create_session(user_id)
                    
                    time_diff = (datetime.now() - last_timestamp).total_seconds() / 60  # 转换为分钟
                    
                    if time_diff > timeout_minutes:
                        # 超过超时时间，创建新session（这是新对话）
                        # session超时，创建新session，不输出日志
                        return self.create_session(user_id)
                    else:
                        # 在超时时间内，使用原session（同一对话的继续）
                        # session有效，继续使用，不输出日志
                        return session_id
                    
            except Exception as e:
                logger.error(f"❌ 验证session失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    try:
                        self.reconnect()
                    except:
                        pass
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    # 如果验证失败，创建新session
                    logger.warning(f"⚠️ 验证session失败，创建新session")
                    return self.create_session(user_id)
        
        # 最终失败时创建新session
        return self.create_session(user_id)
    
    def check_session_exists(self, user_id: str, session_id: str) -> bool:
        """检查session是否存在且属于该用户"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()
                
                with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    # 检查session是否存在：在users表中存在，或者在interactions表中有记录
                    sql = """
                    SELECT COUNT(*) as count
                    FROM (
                        SELECT session_id FROM users WHERE user_id = %s AND session_id = %s
                        UNION
                        SELECT DISTINCT session_id FROM interactions WHERE user_id = %s AND session_id = %s
                    ) as combined
                    """
                    cursor.execute(sql, (user_id, session_id, user_id, session_id))
                    result = cursor.fetchone()
                    exists = result['count'] > 0 if result else False
                    # 检查session，不输出日志
                    return exists
                    
            except Exception as e:
                logger.error(f"❌ 检查session存在失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    try:
                        self.reconnect()
                    except:
                        pass
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    # 如果检查失败，返回False（不存在），这样会创建新session
                    return False
        return False
    
    def end_session(self, session_id: str) -> bool:
        """结束用户会话（带连接健康检查和自动重连）"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # 检查连接健康状态
                if not self.is_connection_healthy():
                    logger.debug(f"数据库连接不健康，尝试重连 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()
                
                with self.connection.cursor() as cursor:
                    sql = """
                    UPDATE users 
                    SET session_id = NULL, app_type = NULL, device_info = NULL, ip_address = NULL
                    WHERE session_id = %s
                    """
                    cursor.execute(sql, (session_id,))
                    self.connection.commit()
                    
                    if cursor.rowcount > 0:
                        # 会话结束成功，不输出日志
                        return True
                    else:
                        # 会话不存在，不一定是错误，可能是已经结束过了
                        return False
            except (pymysql.Error, ConnectionError, OSError) as e:
                error_msg = str(e)
                # 检查是否是连接相关错误
                if any(keyword in error_msg.lower() for keyword in ['gone away', 'lost connection', 'broken pipe', 'connection aborted', '10053']):
                    logger.debug(f"数据库连接错误，尝试重连 (尝试 {attempt + 1}/{max_retries}): {error_msg}")
                    if attempt < max_retries - 1:
                        try:
                            self.reconnect()
                            time.sleep(self.retry_delay * (attempt + 1))
                            continue
                        except Exception as reconnect_error:
                            logger.error(f"重连失败: {reconnect_error}")
                    else:
                        logger.error(f"❌ 结束会话失败（连接错误，已重试 {max_retries} 次）: {error_msg}")
                        return False
                else:
                    # 其他类型的错误，直接返回
                    if attempt == max_retries - 1:
                        logger.error(f"❌ 结束会话失败: {error_msg}")
                    return False
            except Exception as e:
                # 其他未知错误
                if attempt == max_retries - 1:
                    logger.error(f"❌ 结束会话失败: {e}")
                return False
        
        return False
    
    def log_interaction(self, user_id: str, interaction_type: str, content: str, 
                       response: str = None, session_id: str = None, 
                       duration_seconds: int = None, success: bool = True, 
                       error_message: str = None, force_new_session: bool = False) -> tuple[bool, str]:
        """记录交互 - 确保每条记录都关联到session（使用独立连接，避免并发冲突）"""
        max_retries = 3
        connection = None
        for attempt in range(max_retries):
            try:
                # 使用独立连接避免并发冲突
                connection = self._get_fresh_connection()
                
                # 简化逻辑：如果session_id为空，创建新session；如果提供了session_id，直接使用
                if force_new_session or not session_id or session_id.strip() == '':
                    session_id = self.create_session(user_id)
                    if not session_id:
                        logger.error(f"❌ 无法为用户 {user_id} 创建session")
                        if connection:
                            try:
                                connection.close()
                            except:
                                pass
                        return False, None
                    # 创建新session，不输出日志
                else:
                    # 直接使用提供的session_id，不做任何验证或复用检查
                    # 使用提供的session，不输出日志
                    pass
                
                # 获取用户名（禁止使用unknown或None）
                user_info = self.get_user_by_id(user_id)
                if not user_info:
                    logger.error(f"❌ 无法获取用户信息: user_id={user_id}, 无法记录交互")
                    if connection:
                        try:
                            connection.close()
                        except:
                            pass
                    return False, None
                
                username = user_info.get('username')
                if not username or username == 'unknown':
                    logger.error(f"❌ 用户名无效: user_id={user_id}, username={username}, 无法记录交互")
                    if connection:
                        try:
                            connection.close()
                        except:
                            pass
                    return False, None
                
                with connection.cursor() as cursor:
                    sql = """
                    INSERT INTO interactions 
                    (user_id, username, interaction_type, content, response, session_id, 
                     duration_seconds, success, error_message)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    cursor.execute(sql, (user_id, username, interaction_type, content, response, 
                                       session_id, duration_seconds, success, error_message))
                    connection.commit()
                    
                    if connection:
                        connection.close()
                    # 交互记录成功，不输出日志
                    return True, session_id
                    
            except Exception as e:
                # 只记录有意义的错误
                if str(e) and str(e) != "(0, '')":
                    logger.error(f"❌ 记录交互失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    logger.error(f"❌ 记录交互最终失败，已重试 {max_retries} 次")
                    return False, None
                    return False, None
    
    def get_user_interactions(self, user_id: str, limit: int = 50, offset: int = 0) -> List[Dict]:
        """获取用户交互记录"""
        try:
            with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                sql = """
                SELECT * FROM interactions 
                WHERE user_id = %s 
                ORDER BY timestamp DESC 
                LIMIT %s OFFSET %s
                """
                cursor.execute(sql, (user_id, limit, offset))
                return cursor.fetchall()
        except Exception as e:
            logger.error(f"❌ 获取交互记录失败: {e}")
            return []
    
    def get_session_interactions(self, session_id: str, limit: int = 100, offset: int = 0) -> List[Dict]:
        """获取指定session下的所有交互记录"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()
                
                with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    sql = """
                    SELECT * FROM interactions 
                    WHERE session_id = %s 
                    ORDER BY timestamp ASC 
                    LIMIT %s OFFSET %s
                    """
                    cursor.execute(sql, (session_id, limit, offset))
                    results = cursor.fetchall()
                    # 获取session交互记录成功，不输出日志
                    return results
            except Exception as e:
                logger.error(f"❌ 获取session交互记录失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    try:
                        self.reconnect()
                    except:
                        pass
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return []
        return []
    
    def get_interaction_stats(self, user_id: str = None, days: int = 30) -> Dict:
        """获取交互统计"""
        try:
            with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                where_clause = "WHERE timestamp >= DATE_SUB(NOW(), INTERVAL %s DAY)"
                params = [days]
                
                if user_id:
                    where_clause += " AND user_id = %s"
                    params.append(user_id)
                
                sql = f"""
                SELECT 
                    interaction_type,
                    COUNT(*) as count,
                    AVG(duration_seconds) as avg_duration,
                    SUM(CASE WHEN success = 1 THEN 1 ELSE 0 END) as success_count,
                    SUM(CASE WHEN success = 0 THEN 1 ELSE 0 END) as failure_count
                FROM interactions 
                {where_clause}
                GROUP BY interaction_type
                """
                cursor.execute(sql, params)
                stats = cursor.fetchall()
                
                # 计算总体统计
                total_sql = f"""
                SELECT 
                    COUNT(*) as total_interactions,
                    AVG(duration_seconds) as avg_duration,
                    SUM(CASE WHEN success = 1 THEN 1 ELSE 0 END) as total_success,
                    SUM(CASE WHEN success = 0 THEN 1 ELSE 0 END) as total_failure
                FROM interactions 
                {where_clause}
                """
                cursor.execute(total_sql, params)
                total_stats = cursor.fetchone()
                
                return {
                    'by_type': stats,
                    'total': total_stats
                }
        except Exception as e:
            logger.error(f"❌ 获取交互统计失败: {e}")
            return {}
    
    def log_system_event(self, log_level: str, service_name: str, message: str):
        """记录系统日志（使用独立连接，静默失败）"""
        connection = None
        try:
            # 使用独立连接避免并发冲突
            connection = self._get_fresh_connection()
            
            with connection.cursor() as cursor:
                sql = """
                INSERT INTO system_logs 
                (log_level, service_name, message)
                VALUES (%s, %s, %s)
                """
                cursor.execute(sql, (log_level, service_name, message))
                connection.commit()
                
            if connection:
                connection.close()
        except Exception as e:
            # 静默失败，不记录日志（避免日志循环）
            # 只记录有意义的错误
            if str(e) and str(e) != "(0, '')" and "MySQL" not in str(e):
                logger.error(f"❌ 记录系统日志失败: {e}")
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    def get_active_users(self, hours: int = 24) -> List[Dict]:
        """获取活跃用户"""
        try:
            with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                sql = """
                SELECT DISTINCT u.user_id, u.username, u.last_login_at, u.last_logout_at,
                       COUNT(i.id) as interaction_count
                FROM users u
                LEFT JOIN interactions i ON u.user_id = i.user_id 
                    AND i.timestamp >= DATE_SUB(NOW(), INTERVAL %s HOUR)
                WHERE u.last_login_at >= DATE_SUB(NOW(), INTERVAL %s HOUR)
                GROUP BY u.user_id, u.username, u.last_login_at, u.last_logout_at
                ORDER BY u.last_login_at DESC
                """
                cursor.execute(sql, (hours, hours))
                return cursor.fetchall()
        except Exception as e:
            logger.error(f"❌ 获取活跃用户失败: {e}")
            return []
    
    # 移除TTS相关函数 - 不再需要TTS播放计数和时间字段
    
    def get_tts_stats(self, user_id: str = None, days: int = 30) -> Dict:
        """获取TTS播放统计"""
        try:
            with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                where_clause = "WHERE timestamp >= DATE_SUB(NOW(), INTERVAL %s DAY)"
                params = [days]
                
                if user_id:
                    where_clause += " AND user_id = %s"
                    params.append(user_id)
                
                sql = f"""
                SELECT 
                    COUNT(*) as total_interactions,
                    SUM(tts_play_count) as total_tts_plays,
                    AVG(tts_play_count) as avg_tts_plays_per_interaction,
                    MAX(tts_play_count) as max_tts_plays,
                    COUNT(CASE WHEN tts_play_count > 0 THEN 1 END) as interactions_with_tts,
                    COUNT(CASE WHEN tts_play_count = 0 THEN 1 END) as interactions_without_tts
                FROM interactions 
                {where_clause}
                """
                cursor.execute(sql, params)
                result = cursor.fetchone()
                
                # 计算TTS播放率
                if result['total_interactions'] > 0:
                    result['tts_play_rate'] = result['interactions_with_tts'] / result['total_interactions']
                else:
                    result['tts_play_rate'] = 0
                
                return result
        except Exception as e:
            logger.error(f"❌ 获取TTS统计失败: {e}")
            return {}
    
    def get_most_played_interactions(self, user_id: str = None, limit: int = 10) -> List[Dict]:
        """获取播放次数最多的交互记录"""
        try:
            with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                where_clause = "WHERE tts_play_count > 0"
                params = []
                
                if user_id:
                    where_clause += " AND user_id = %s"
                    params.append(user_id)
                
                params.append(limit)
                
                sql = f"""
                SELECT 
                    id, user_id, interaction_type, content, response,
                    tts_play_count, last_tts_play_time, timestamp
                FROM interactions 
                {where_clause}
                ORDER BY tts_play_count DESC, last_tts_play_time DESC
                LIMIT %s
                """
                cursor.execute(sql, params)
                return cursor.fetchall()
        except Exception as e:
            logger.error(f"❌ 获取最常播放交互失败: {e}")
            return []
    
    def cleanup_old_data(self, days: int = 90):
        """清理旧数据"""
        try:
            with self.connection.cursor() as cursor:
                # 清理旧的交互记录
                cursor.execute("""
                    DELETE FROM interactions 
                    WHERE timestamp < DATE_SUB(NOW(), INTERVAL %s DAY)
                """, (days,))
                
                # 清理旧的系统日志
                cursor.execute("""
                    DELETE FROM system_logs 
                    WHERE timestamp < DATE_SUB(NOW(), INTERVAL %s DAY)
                """, (days,))
                
                # 清理旧的会话信息（基于最后登录时间）
                cursor.execute("""
                    UPDATE users 
                    SET session_id = NULL, app_type = NULL, device_info = NULL, ip_address = NULL
                    WHERE last_login_at IS NOT NULL 
                    AND last_login_at < DATE_SUB(NOW(), INTERVAL %s DAY)
                    AND session_id IS NOT NULL
                """, (days,))
                
                self.connection.commit()
                # 清理旧数据完成，不输出日志
        except Exception as e:
            logger.error(f"❌ 清理旧数据失败: {e}")
    
    def query_interactions(self, interaction_type: str = None, user_id: str = None, limit: int = 10):
        """查询交互记录"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()

                with self.connection.cursor() as cursor:
                    # 构建查询条件
                    conditions = []
                    params = []
                    
                    if interaction_type:
                        conditions.append("interaction_type = %s")
                        params.append(interaction_type)
                    
                    if user_id:
                        conditions.append("user_id = %s")
                        params.append(user_id)
                    
                    where_clause = ""
                    if conditions:
                        where_clause = "WHERE " + " AND ".join(conditions)
                    
                    sql = f"""
                    SELECT id, user_id, interaction_type, content, response, 
                           timestamp, session_id, duration_seconds, success, error_message
                    FROM interactions 
                    {where_clause}
                    ORDER BY timestamp DESC 
                    LIMIT %s
                    """
                    params.append(limit)
                    
                    cursor.execute(sql, params)
                    results = cursor.fetchall()
                    
                    # 转换为字典列表
                    records = []
                    for row in results:
                        record = {
                            'id': row[0],
                            'user_id': row[1],
                            'interaction_type': row[2],
                            'content': row[3],
                            'response': row[4],
                            'timestamp': row[5].isoformat() if row[5] else None,
                            'session_id': row[6],
                            'duration_seconds': row[7],
                            'success': bool(row[8]),
                            'error_message': row[9]
                        }
                        records.append(record)
                    
                    # 查询交互记录成功，不输出日志
                    return records

            except Exception as e:
                logger.error(f"❌ 查询交互记录失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    self.reconnect()
                    time.sleep(1)
                else:
                    return []
        return []

    # ==================== 故事控制相关功能 ====================
    
    def complete_reading(self, user_id: str, story_id: str, story_title: str, 
                        completion_mode: str, device_info: str = None, username: str = None) -> bool:
        """
        标记故事为已完成
        重要：此方法只能由用户点击完成按钮触发，不能自动调用
        用户点击完成按钮是标记故事完成阅读的唯一标准，没有之一
        禁止使用unknown作为用户名，如果无法获取用户名则返回False
        """
        max_retries = 3
        connection = None
        for attempt in range(max_retries):
            try:
                # 使用独立连接避免并发冲突
                connection = self._get_fresh_connection()

                with connection.cursor() as cursor:
                    # 验证完成方式
                    valid_modes = ['text', 'audio']
                    if completion_mode not in valid_modes:
                        logger.error(f"❌ 无效的完成方式: {completion_mode}")
                        return False
                    
                    # 如果未提供username，从数据库获取（禁止使用unknown）
                    if not username:
                        user_info = self.get_user_by_id(user_id)
                        if not user_info:
                            logger.error(f"❌ 无法获取用户信息: user_id={user_id}")
                            return False
                        username = user_info.get('username')
                        if not username or username == 'unknown':
                            logger.error(f"❌ 用户名无效: user_id={user_id}, username={username}")
                            return False
                    
                    # 检查是否已存在记录
                    check_sql = "SELECT id, is_completed FROM reading_progress WHERE user_id = %s AND story_id = %s"
                    cursor.execute(check_sql, (user_id, story_id))
                    existing = cursor.fetchone()
                    
                    if existing:
                        # 更新现有记录（包括story_title，确保标题信息是最新的）
                        # 保留现有的 current_position 和 total_length，只更新完成状态
                        update_sql = """
                        UPDATE reading_progress 
                        SET story_title = %s, is_completed = TRUE, completion_time = NOW(), 
                            completion_mode = %s, last_read_time = NOW(),
                            reading_progress = CASE 
                                WHEN total_length > 0 THEN (current_position / total_length * 100)
                                ELSE 100.0
                            END,
                            username = %s
                        WHERE user_id = %s AND story_id = %s
                        """
                        cursor.execute(update_sql, (
                            story_title, completion_mode, username, user_id, story_id
                        ))
                    else:
                        # 创建新记录
                        insert_sql = """
                        INSERT INTO reading_progress 
                        (user_id, username, story_id, story_title, current_position, total_length, 
                         reading_progress, is_completed, completion_mode, start_time, completion_time)
                        VALUES (%s, %s, %s, %s, 0, 0, 100.0, TRUE, %s, NOW(), NOW())
                        """
                        cursor.execute(insert_sql, (
                            user_id, username, story_id, story_title, completion_mode
                        ))
                    
                    connection.commit()
                    logger.info(f"✅ 标记故事完成: user_id={user_id}, story_id={story_id}, completion_mode={completion_mode}")
                    if connection:
                        connection.close()
                    return True

            except Exception as e:
                logger.error(f"❌ 标记故事完成失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return False
        return False

    def update_reading_progress(self, user_id: str, story_id: str, story_title: str,
                              current_position: int, total_length: int, 
                              device_info: str = None, username: str = None, 
                              completion_mode: str = None) -> bool:
        """
        更新阅读进度（使用独立连接以避免并发冲突）
        禁止使用unknown作为用户名，如果无法获取用户名则返回False
        """
        max_retries = 3
        connection = None
        for attempt in range(max_retries):
            try:
                # 使用独立连接避免并发冲突
                connection = self._get_fresh_connection()
                
                # 如果未提供username，从数据库获取（禁止使用unknown）
                if not username:
                    user_info = self.get_user_by_id(user_id)
                    if not user_info:
                        logger.error(f"❌ 无法获取用户信息: user_id={user_id}")
                        return False
                    username = user_info.get('username')
                    if not username or username == 'unknown':
                        logger.error(f"❌ 用户名无效: user_id={user_id}, username={username}")
                        return False

                with connection.cursor() as cursor:
                    # 计算阅读进度百分比
                    reading_progress = (current_position / total_length * 100) if total_length > 0 else 0
                    # 不自动设置完成状态，只有通过完成阅读API才能设置
                    is_completed = False
                    
                    # 检查是否已存在记录
                    check_sql = "SELECT id, start_time FROM reading_progress WHERE user_id = %s AND story_id = %s"
                    cursor.execute(check_sql, (user_id, story_id))
                    existing = cursor.fetchone()
                    
                    if existing:
                        # 更新现有记录（包括story_title，确保标题信息是最新的）
                        update_sql = """
                        UPDATE reading_progress 
                        SET story_title = %s, current_position = %s, total_length = %s, reading_progress = %s,
                            is_completed = %s, last_read_time = NOW(),
                            completion_time = CASE WHEN %s = 1 THEN NOW() ELSE completion_time END,
                            completion_mode = CASE WHEN %s = 1 AND %s IS NOT NULL THEN %s ELSE completion_mode END,
                            username = %s
                        WHERE user_id = %s AND story_id = %s
                        """
                        cursor.execute(update_sql, (
                            story_title, current_position, total_length, reading_progress, is_completed,
                            is_completed, is_completed, completion_mode, completion_mode,
                            username, user_id, story_id
                        ))
                    else:
                        # 创建新记录
                        insert_sql = """
                        INSERT INTO reading_progress 
                        (user_id, username, story_id, story_title, current_position, total_length, 
                         reading_progress, is_completed, completion_mode, start_time, completion_time)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), 
                                CASE WHEN %s = 1 THEN NOW() ELSE NULL END)
                        """
                        cursor.execute(insert_sql, (
                            user_id, username, story_id, story_title, current_position, total_length,
                            reading_progress, is_completed, completion_mode, is_completed
                        ))
                    
                    connection.commit()
                    # 更新阅读进度成功，不输出日志
                    if connection:
                        connection.close()
                    return True

            except Exception as e:
                logger.error(f"❌ 更新阅读进度失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    self.reconnect()
                    time.sleep(1)
                else:
                    return False
        return False

    def get_reading_progress(self, user_id: str, story_id: str = None) -> List[Dict[str, Any]]:
        """获取阅读进度"""
        max_retries = 3
        for attempt in range(max_retries):
            connection = None
            try:
                # 使用新的连接避免连接状态问题
                connection = self._get_fresh_connection()

                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    if story_id:
                        # 获取特定故事的进度
                        sql = """
                        SELECT story_id, story_title, current_position, total_length, 
                               reading_progress, is_completed, start_time, last_read_time, 
                               completion_time
                        FROM reading_progress 
                        WHERE user_id = %s AND story_id = %s
                        ORDER BY last_read_time DESC
                        """
                        cursor.execute(sql, (user_id, story_id))
                    else:
                        # 获取用户所有故事的进度
                        sql = """
                        SELECT story_id, story_title, current_position, total_length, 
                               reading_progress, is_completed, start_time, last_read_time, 
                               completion_time
                        FROM reading_progress 
                        WHERE user_id = %s
                        ORDER BY last_read_time DESC
                        """
                        cursor.execute(sql, (user_id,))
                    
                    results = cursor.fetchall()
                    
                    # 处理结果
                    progress_list = []
                    for row in results:
                        progress = {
                            'story_id': row['story_id'],
                            'story_title': row['story_title'],
                            'current_position': row['current_position'],
                            'total_length': row['total_length'],
                            'reading_progress': float(row['reading_progress']) if row['reading_progress'] else 0.0,
                            'is_completed': bool(row['is_completed']),
                            'start_time': row['start_time'].isoformat() if row['start_time'] else None,
                            'last_read_time': row['last_read_time'].isoformat() if row['last_read_time'] else None,
                            'completion_time': row['completion_time'].isoformat() if row['completion_time'] else None
                        }
                        progress_list.append(progress)
                    
                    # 获取阅读进度成功，不输出日志
                    return progress_list

            except Exception as e:
                logger.error(f"❌ 获取阅读进度失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return []
            finally:
                if connection:
                    connection.close()
        return []

    def log_story_interaction(self, user_id: str, story_id: str, interaction_type: str,
                            interaction_data: Dict[str, Any] = None, device_info: str = None) -> bool:
        """记录故事交互（app_version字段已删除）"""
        max_retries = 3
        for attempt in range(max_retries):
            connection = None
            try:
                # 使用新的连接避免连接状态问题
                connection = self._get_fresh_connection()
                
                # 获取用户名（禁止使用unknown或None）
                user_info = self.get_user_by_id(user_id)
                if not user_info:
                    logger.error(f"❌ 无法获取用户信息: user_id={user_id}, 无法记录故事交互")
                    return False
                
                username = user_info.get('username')
                if not username or username == 'unknown':
                    logger.error(f"❌ 用户名无效: user_id={user_id}, username={username}, 无法记录故事交互")
                    return False
                
                with connection.cursor() as cursor:
                    sql = """
                    INSERT INTO story_interactions 
                    (user_id, username, story_id, interaction_type, interaction_data, device_info)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """
                    interaction_json = json.dumps(interaction_data) if interaction_data else None
                    cursor.execute(sql, (user_id, username, story_id, interaction_type, interaction_json, device_info))
                    connection.commit()
                    
                    logger.info(f"记录故事交互成功: {user_id} ({username}) - {story_id} - {interaction_type}")
                    return True

            except Exception as e:
                logger.error(f"❌ 记录故事交互失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return False
            finally:
                if connection:
                    connection.close()
        return False

    def get_reading_statistics(self, user_id: str, days: int = 30) -> Dict[str, Any]:
        """获取阅读统计"""
        max_retries = 3
        for attempt in range(max_retries):
            connection = None
            try:
                # 使用新的连接避免连接状态问题
                connection = self._get_fresh_connection()

                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    # 获取基本统计
                    stats_sql = """
                    SELECT 
                        COUNT(DISTINCT story_id) as total_stories,
                        COUNT(CASE WHEN is_completed = TRUE THEN 1 END) as completed_stories,
                        AVG(reading_progress) as avg_progress,
                        MAX(last_read_time) as last_reading_time
                    FROM reading_progress 
                    WHERE user_id = %s AND last_read_time >= DATE_SUB(NOW(), INTERVAL %s DAY)
                    """
                    cursor.execute(stats_sql, (user_id, days))
                    stats_result = cursor.fetchone()
                    
                    # 获取最近阅读的故事
                    recent_sql = """
                    SELECT story_id, story_title, reading_progress, is_completed, last_read_time
                    FROM reading_progress 
                    WHERE user_id = %s
                    ORDER BY last_read_time DESC
                    LIMIT 10
                    """
                    cursor.execute(recent_sql, (user_id,))
                    recent_stories = cursor.fetchall()
                    
                    # 获取每日阅读统计（已移除reading_duration_seconds字段）
                    daily_sql = """
                    SELECT DATE(last_read_time) as reading_date, 
                           COUNT(*) as daily_count
                    FROM reading_progress 
                    WHERE user_id = %s AND last_read_time >= DATE_SUB(NOW(), INTERVAL %s DAY)
                    GROUP BY DATE(last_read_time)
                    ORDER BY reading_date DESC
                    """
                    cursor.execute(daily_sql, (user_id, days))
                    daily_stats = cursor.fetchall()
                    
                    # 构建统计结果
                    statistics = {
                        'total_stories': stats_result['total_stories'] or 0,
                        'completed_stories': stats_result['completed_stories'] or 0,
                        'total_reading_time_seconds': 0,  # Field removed (reading_duration_seconds)
                        'average_progress': float(stats_result['avg_progress']) if stats_result['avg_progress'] else 0.0,
                        'last_reading_time': stats_result['last_reading_time'].isoformat() if stats_result['last_reading_time'] else None,
                        'recent_stories': [],
                        'daily_reading': []
                    }
                    
                    # 处理最近阅读的故事
                    for story in recent_stories:
                        statistics['recent_stories'].append({
                            'story_id': story['story_id'],
                            'story_title': story['story_title'],
                            'reading_progress': float(story['reading_progress']) if story['reading_progress'] else 0.0,
                            'is_completed': bool(story['is_completed']),
                            'last_read_time': story['last_read_time'].isoformat() if story['last_read_time'] else None
                        })
                    
                    # 处理每日阅读统计
                    for daily in daily_stats:
                        statistics['daily_reading'].append({
                            'date': daily['reading_date'].isoformat() if daily['reading_date'] else None,
                            'count': daily['daily_count'] or 0
                        })
                    
                    # 获取阅读统计成功，不输出日志
                    return statistics

            except Exception as e:
                logger.error(f"❌ 获取阅读统计失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return {}
            finally:
                if connection:
                    connection.close()
        return {}

    def get_all_users_reading_progress(self, limit=100, offset=0):
        """获取所有用户的阅读进度（管理员功能）"""
        max_retries = 3
        for attempt in range(max_retries):
            connection = None
            try:
                # 使用新的连接避免连接状态问题
                connection = self._get_fresh_connection()

                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    sql = """
                    SELECT rp.*, u.username 
                    FROM reading_progress rp
                    LEFT JOIN users u ON rp.user_id = u.user_id
                    WHERE rp.id IN (
                        SELECT MAX(id) 
                        FROM reading_progress 
                        GROUP BY user_id, story_id
                    )
                    ORDER BY rp.last_read_time DESC
                    LIMIT %s OFFSET %s
                    """
                    cursor.execute(sql, (limit, offset))
                    results = cursor.fetchall()
                    
                    # 获取总数（只统计唯一的用户-故事组合）
                    count_sql = """
                    SELECT COUNT(*) as count FROM (
                        SELECT MAX(id) 
                        FROM reading_progress 
                        GROUP BY user_id, story_id
                    ) as unique_records
                    """
                    cursor.execute(count_sql)
                    total_count = cursor.fetchone()['count']
                    
                    return {
                        'progress_list': results,
                        'total_count': total_count,
                        'limit': limit,
                        'offset': offset
                    }

            except Exception as e:
                logger.error(f"❌ 获取所有用户阅读进度失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return None
            finally:
                if connection:
                    connection.close()
        return None

    def delete_reading_record(self, record_id: int) -> bool:
        """删除阅读记录"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()

                with self.connection.cursor() as cursor:
                    # 删除指定的阅读记录
                    delete_sql = "DELETE FROM reading_progress WHERE id = %s"
                    cursor.execute(delete_sql, (record_id,))
                    
                    if cursor.rowcount > 0:
                        self.connection.commit()
                        # 删除阅读记录成功，不输出日志
                        return True
                    else:
                        logger.warning(f"⚠️ 未找到要删除的记录: ID={record_id}")
                        return False

            except Exception as e:
                logger.error(f"❌ 删除阅读记录失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    self.reconnect()
                    time.sleep(1)
                else:
                    return False
        return False

    def reset_user_password(self, user_id: str, new_password: str) -> bool:
        """重置用户密码"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()

                with self.connection.cursor() as cursor:
                    # 生成密码哈希
                    import hashlib
                    password_hash = hashlib.sha256(new_password.encode()).hexdigest()
                    
                    # 更新用户密码（同时保存原始密码和哈希值）
                    update_sql = "UPDATE users SET password_hash = %s, original_password = %s WHERE user_id = %s"
                    cursor.execute(update_sql, (password_hash, new_password, user_id))
                    
                    if cursor.rowcount > 0:
                        self.connection.commit()
                        # 重置用户密码成功，不输出日志
                        return True
                    else:
                        logger.warning(f"⚠️ 未找到要重置密码的用户: {user_id}")
                        return False

            except Exception as e:
                logger.error(f"❌ 重置用户密码失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    self.reconnect()
                    time.sleep(1)
                else:
                    return False
        return False

    def get_user_password_info(self, user_id: str) -> dict:
        """获取用户密码信息"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()

                with self.connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    # 获取用户密码信息
                    sql = """
                    SELECT user_id, username, password_hash, original_password, created_at, last_login_at
                    FROM users 
                    WHERE user_id = %s
                    """
                    cursor.execute(sql, (user_id,))
                    result = cursor.fetchone()
                    
                    if result:
                        # 返回密码信息，包括原始密码（用于管理员查看）
                        return {
                            'user_id': result['user_id'],
                            'username': result['username'],
                            'has_password': bool(result['password_hash']),
                            'password': result.get('original_password', '未设置'),
                            'password_set_date': result['created_at'],
                            'last_login': result['last_login_at']
                        }
                    else:
                        return None

            except Exception as e:
                logger.error(f"❌ 获取用户密码信息失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    self.reconnect()
                    time.sleep(1)
                else:
                    return None
        return None

    def admin_update_reading_completion(self, user_id, story_id, is_completed, admin_user_id):
        """管理员更新用户阅读完成状态"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()

                with self.connection.cursor() as cursor:
                    # 检查记录是否存在
                    check_sql = "SELECT id, is_completed FROM reading_progress WHERE user_id = %s AND story_id = %s"
                    cursor.execute(check_sql, (user_id, story_id))
                    existing = cursor.fetchone()
                    
                    if not existing:
                        return False, "阅读记录不存在"
                    
                    record_id, current_status = existing
                    
                    # 更新完成状态
                    if is_completed:
                        update_sql = """
                        UPDATE reading_progress 
                        SET is_completed = 1, completion_time = NOW(), last_read_time = NOW()
                        WHERE user_id = %s AND story_id = %s
                        """
                        cursor.execute(update_sql, (user_id, story_id))
                        
                        # 管理员操作记录已移除（admin_operations表已删除）
                    else:
                        update_sql = """
                        UPDATE reading_progress 
                        SET is_completed = 0, completion_time = NULL, last_read_time = NOW()
                        WHERE user_id = %s AND story_id = %s
                        """
                        cursor.execute(update_sql, (user_id, story_id))
                        
                        # 管理员操作记录已移除（admin_operations表已删除）
                    
                    self.connection.commit()
                    
                    action = "标记为已完成" if is_completed else "取消完成状态"
                    # 管理员操作成功，不输出日志
                    return True, f"成功{action}"

            except Exception as e:
                logger.error(f"❌ 管理员更新阅读完成状态失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    self.reconnect()
                    time.sleep(1)
                else:
                    return False, str(e)
        return False, "操作失败"

    # log_admin_operation 方法已删除（admin_operations表已删除）

    def get_user_by_id(self, user_id):
        """根据用户ID获取用户基本信息"""
        max_retries = 3
        for attempt in range(max_retries):
            connection = None
            try:
                # 使用新的连接避免连接状态问题
                connection = self._get_fresh_connection()

                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    sql = """
                    SELECT user_id, username, created_at, last_login_at, is_active
                    FROM users 
                    WHERE user_id = %s
                    """
                    cursor.execute(sql, (user_id,))
                    result = cursor.fetchone()
                    
                    if connection:
                        connection.close()
                    return result

            except Exception as e:
                # 只记录有意义的错误
                if str(e) and str(e) != "(0, '')":
                    logger.error(f"❌ 获取用户信息失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return None
            finally:
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
        return None

    def get_user_reading_progress_details(self, user_id):
        """获取用户阅读进度详情"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.connection or not self.connection.open:
                    logger.warning(f"⚠️ 数据库连接已关闭，尝试重新连接 (尝试 {attempt + 1}/{max_retries})")
                    self.reconnect()

                with self.connection.cursor() as cursor:
                    sql = """
                    SELECT rp.story_id, rp.story_title, rp.current_position, rp.total_length,
                           rp.reading_progress, rp.is_completed, rp.start_time, rp.last_read_time,
                           rp.completion_time
                    FROM reading_progress rp
                    WHERE rp.user_id = %s
                    ORDER BY rp.last_read_time DESC
                    """
                    cursor.execute(sql, (user_id,))
                    columns = [desc[0] for desc in cursor.description]
                    progress_list = []
                    
                    for row in cursor.fetchall():
                        progress = dict(zip(columns, row))
                        progress_list.append(progress)
                    
                    return progress_list

            except Exception as e:
                logger.error(f"❌ 获取用户阅读进度详情失败: {e}")
                if attempt < max_retries - 1:
                    self.reconnect()
                    time.sleep(1)
                else:
                    return []
        return []

    def get_user_reading_summary(self, user_id):
        """获取用户阅读摘要（管理员查看）"""
        max_retries = 3
        for attempt in range(max_retries):
            connection = None
            try:
                # 使用新的连接避免连接状态问题
                connection = self._get_fresh_connection()
                
                with connection.cursor() as cursor:
                    # 获取用户基本信息
                    user_sql = "SELECT username, created_at, last_login_at FROM users WHERE user_id = %s"
                    cursor.execute(user_sql, (user_id,))
                    user_info = cursor.fetchone()
                    
                    if not user_info:
                        return None
                    
                    # 获取阅读统计
                    stats_sql = """
                    SELECT 
                        COUNT(*) as total_stories,
                        SUM(CASE WHEN is_completed = 1 THEN 1 ELSE 0 END) as completed_stories,
                        AVG(reading_progress) as avg_progress,
                        MAX(last_read_time) as last_reading_time
                    FROM reading_progress 
                    WHERE user_id = %s
                    """
                    cursor.execute(stats_sql, (user_id,))
                    stats = cursor.fetchone()
                    
                    return {
                        'user_id': user_id,
                        'username': user_info[0],
                        'created_at': user_info[1],
                        'last_login_at': user_info[2],
                        'total_stories': stats[0] or 0,
                        'completed_stories': stats[1] or 0,
                        'avg_progress': float(stats[2]) if stats[2] else 0.0,
                        'last_reading_time': stats[3]
                    }

            except Exception as e:
                logger.error(f"❌ 获取用户阅读摘要失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return None
            finally:
                if connection:
                    connection.close()
        return None

    # ==================== 故事管理方法 ====================
    
    def get_all_stories(self, include_inactive: bool = False) -> List[Dict]:
        """
        获取所有故事列表
        注意：stories表已删除，故事数据现在从Excel文件读取（只返回30个故事）
        """
        import os
        from openpyxl import load_workbook
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # 项目根目录
                project_root = os.path.dirname(__file__)
                
                # Excel文件路径
                excel_file = os.path.join(project_root, 'Story_v2.xlsx')
                
                # 故事音频文件目录
                story_audio_dir = os.path.join(
                    project_root,
                    'story_control_app', 'app', 'src', 'main', 'assets', 'story_audio'
                )
                
                stories = []
                
                # 检查Excel文件是否存在
                if not os.path.exists(excel_file):
                    logger.warning(f"⚠️ Excel文件不存在: {excel_file}")
                    return []
                
                # 读取Excel文件
                wb = load_workbook(excel_file, rich_text=True)
                ws = wb.active
                
                # 获取列名
                headers = [cell.value for cell in ws[1]]
                
                # 查找标题和内容列
                title_col = None
                content_col = None
                for h in headers:
                    if h and ("题目" in str(h) or "标题" in str(h) or "title" in str(h).lower()):
                        title_col = h
                    if h and ("内容" in str(h) or "content" in str(h).lower()):
                        content_col = h
                
                if not title_col:
                    logger.error("❌ 无法找到标题列")
                    return []
                
                # 读取所有故事数据（只读取30个）
                excel_stories = []
                for row_idx in range(2, min(ws.max_row + 1, 32)):  # 最多31行（1行标题+30行数据）
                    story = {}
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        
                        # 处理内容列（富文本）
                        if header == content_col and value:
                            content_parts = []
                            if isinstance(value, CellRichText):
                                for block in value:
                                    if isinstance(block, TextBlock):
                                        text = block.text
                                        is_bold = False
                                        if block.font and hasattr(block.font, 'b') and block.font.b:
                                            is_bold = True
                                        content_parts.append({"text": text, "bold": is_bold})
                                    else:
                                        content_parts.append({"text": str(block), "bold": False})
                            else:
                                text = str(value) if value else ""
                                is_bold = False
                                if cell.font and hasattr(cell.font, 'b') and cell.font.b:
                                    is_bold = True
                                content_parts.append({"text": text, "bold": is_bold})
                            story[header] = content_parts
                        else:
                            story[header] = value
                    
                    # 只添加非空故事
                    if any(v for v in story.values() if v):
                        excel_stories.append(story)
                
                logger.info(f"📚 从Excel读取到 {len(excel_stories)} 个故事")
                
                # 获取音频文件列表（用于匹配）
                audio_files = {}
                if os.path.exists(story_audio_dir):
                    for f in os.listdir(story_audio_dir):
                        if f.endswith('.mp3'):
                            # 去掉扩展名作为key
                            audio_key = f.replace('.mp3', '')
                            audio_files[audio_key] = f
                
                # 为每个Excel中的故事创建故事对象
                for idx, excel_story in enumerate(excel_stories, 1):
                    title = excel_story.get(title_col, "")
                    if not title:
                        continue
                    
                    # 处理内容（完整保留所有文本，包括换行符）
                    content_parts = excel_story.get(content_col, [])
                    content_str = ""
                    if isinstance(content_parts, list):
                        for part in content_parts:
                            text = part.get("text", "") if isinstance(part, dict) else str(part)
                            is_bold = part.get("bold", False) if isinstance(part, dict) else False
                            if text:  # 只处理非空文本
                                if is_bold:
                                    content_str += f"**{text}**"
                                else:
                                    content_str += text
                    else:
                        # 如果不是列表，直接转换为字符串（保留换行符）
                        content_str = str(content_parts) if content_parts else ""
                    
                    # 确保内容不为空，如果为空则使用标题作为占位符
                    if not content_str or content_str.strip() == "":
                        content_str = title
                        logger.warning(f"⚠️ 故事 '{title}' 的内容为空，使用标题作为占位符")
                    
                    # 记录内容长度用于调试
                    logger.debug(f"📝 故事 '{title}' 内容长度: {len(content_str)} 字符")
                    
                    # 匹配音频文件（尝试精确匹配和模糊匹配）
                    audio_file = None
                    audio_file_path = None
                    
                    # 精确匹配
                    if title in audio_files:
                        audio_file = audio_files[title]
                    else:
                        # 模糊匹配：查找包含标题关键字的音频文件
                        title_keywords = title.replace('"', '').replace('：', ':').replace('，', ',')
                        for audio_key, audio_name in audio_files.items():
                            # 检查标题是否在音频文件名中，或音频文件名是否在标题中
                            if title_keywords in audio_key or audio_key in title_keywords:
                                audio_file = audio_name
                                break
                    
                    if audio_file:
                        audio_file_path = f"story_audio/{audio_file}"
                    else:
                        logger.warning(f"⚠️ 未找到故事 '{title}' 的音频文件")
                    
                    # 生成故事ID（基于索引，1-30）
                    story_id = f"story_{idx:03d}"
                    
                    story = {
                        'story_id': story_id,
                        'title': title,
                        'content': content_str,
                        'audio_file_path': audio_file_path,
                        'audio_duration_seconds': 0,
                        'is_active': True,
                        'created_at': None,
                        'updated_at': None
                    }
                    stories.append(story)
                
                logger.info(f"✅ 成功生成 {len(stories)} 个故事对象（来自Excel）")
                return stories
                
            except Exception as e:
                logger.error(f"❌ 获取故事列表失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                import traceback
                logger.error(traceback.format_exc())
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    return []
        return []

    def close(self):
        """关闭数据库连接"""
        if self.connection:
            self.connection.close()
            # 数据库连接已关闭，不输出日志
    
    # error_reports 表相关方法已删除（error_reports表已删除）
    def report_error(self, user_id: str = None, username: str = None, error_type: str = 'unknown',
                     error_level: str = 'ERROR', error_message: str = '', error_stack: str = None,
                     error_context: dict = None, app_version: str = None, device_info: str = None,
                     os_version: str = None, screen_info: str = None, network_type: str = None,
                     session_id: str = None, api_endpoint: str = None, request_data: str = None,
                     response_data: str = None) -> bool:
        """报告错误（已禁用，error_reports表已删除）"""
        # error_reports 表已删除，此方法不再执行任何操作
        return True
        max_retries = 3
        for attempt in range(max_retries):
            try:
                connection = self._get_fresh_connection()
                
                with connection.cursor() as cursor:
                    # 检查是否存在相同的错误（基于错误类型和消息的前100个字符）
                    error_signature = error_message[:100] if error_message else ''
                    
                    # 尝试查找相同的错误
                    check_sql = """
                    SELECT id, occurrence_count FROM error_reports 
                    WHERE error_type = %s AND error_message LIKE %s 
                    AND is_resolved = FALSE
                    ORDER BY last_occurrence DESC LIMIT 1
                    """
                    cursor.execute(check_sql, (error_type, f"{error_signature}%"))
                    existing = cursor.fetchone()
                    
                    if existing:
                        # 更新现有错误记录
                        update_sql = """
                        UPDATE error_reports 
                        SET occurrence_count = occurrence_count + 1,
                            last_occurrence = CURRENT_TIMESTAMP,
                            user_id = COALESCE(%s, user_id),
                            username = COALESCE(%s, username),
                            error_context = COALESCE(%s, error_context),
                            device_info = COALESCE(%s, device_info),
                            session_id = COALESCE(%s, session_id)
                        WHERE id = %s
                        """
                        context_json = json.dumps(error_context, ensure_ascii=False) if error_context else None
                        cursor.execute(update_sql, (
                            user_id, username, context_json, device_info, session_id, existing['id']
                        ))
                    else:
                        # 插入新错误记录
                        insert_sql = """
                        INSERT INTO error_reports 
                        (user_id, username, error_type, error_level, error_message, error_stack,
                         error_context, app_version, device_info, os_version, screen_info,
                         network_type, session_id, api_endpoint, request_data, response_data)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """
                        context_json = json.dumps(error_context, ensure_ascii=False) if error_context else None
                        cursor.execute(insert_sql, (
                            user_id, username, error_type, error_level, error_message, error_stack,
                            context_json, app_version, device_info, os_version, screen_info,
                            network_type, session_id, api_endpoint, request_data, response_data
                        ))
                    
                    connection.commit()
                
                if connection:
                    connection.close()
                return True
                    
            except Exception as e:
                # 静默失败，避免日志循环
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return False
        return False
    
    def get_error_reports(self, user_id: str = None, error_type: str = None, 
                         error_level: str = None, is_resolved: bool = None,
                         limit: int = 100, offset: int = 0) -> List[Dict]:
        """获取错误报告列表（已禁用，error_reports表已删除）"""
        # error_reports 表已删除，返回空列表
        return []
        max_retries = 3
        for attempt in range(max_retries):
            try:
                connection = self._get_fresh_connection()
                
                with connection.cursor(pymysql.cursors.DictCursor) as cursor:
                    sql = "SELECT * FROM error_reports WHERE 1=1"
                    params = []
                    
                    if user_id:
                        sql += " AND user_id = %s"
                        params.append(user_id)
                    if error_type:
                        sql += " AND error_type = %s"
                        params.append(error_type)
                    if error_level:
                        sql += " AND error_level = %s"
                        params.append(error_level)
                    if is_resolved is not None:
                        sql += " AND is_resolved = %s"
                        params.append(is_resolved)
                    
                    sql += " ORDER BY last_occurrence DESC LIMIT %s OFFSET %s"
                    params.extend([limit, offset])
                    
                    cursor.execute(sql, params)
                    results = cursor.fetchall()
                    
                    # 转换JSON字段
                    for result in results:
                        if result.get('error_context'):
                            try:
                                result['error_context'] = json.loads(result['error_context'])
                            except:
                                result['error_context'] = {}
                    
                if connection:
                    connection.close()
                return results
                    
            except Exception as e:
                logger.error(f"❌ 获取错误报告失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return []
        return []
    
    def resolve_error(self, error_id: int, resolved_by: str, resolution_note: str = None) -> bool:
        """标记错误为已解决（已禁用，error_reports表已删除）"""
        # error_reports 表已删除，此方法不再执行任何操作
        return True
        max_retries = 3
        for attempt in range(max_retries):
            try:
                connection = self._get_fresh_connection()
                
                with connection.cursor() as cursor:
                    sql = """
                    UPDATE error_reports 
                    SET is_resolved = TRUE,
                        resolved_at = CURRENT_TIMESTAMP,
                        resolved_by = %s,
                        resolution_note = %s
                    WHERE id = %s
                    """
                    cursor.execute(sql, (resolved_by, resolution_note, error_id))
                    connection.commit()
                
                if connection:
                    connection.close()
                return True
                    
            except Exception as e:
                logger.error(f"❌ 标记错误已解决失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if connection:
                    try:
                        connection.close()
                    except:
                        pass
                if attempt < max_retries - 1:
                    time.sleep(self.retry_delay * (attempt + 1))
                else:
                    return False
        return False

# 全局数据库管理器实例
db_manager = DatabaseManager()
